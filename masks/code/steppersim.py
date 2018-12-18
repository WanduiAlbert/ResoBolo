from __future__ import division

import gdspy
import numpy as np
import openpyxl
import argparse
import sys
import os
import pdb

allowed_element_types = set([gdspy.Cell, gdspy.CellReference, gdspy.CellArray])
mask_width = 22000
mask_length = 26000
# Set the library to write all the cells in
main_lib = gdspy.GdsLibrary('main')
gdspy.current_library = main_lib

parser = argparse.ArgumentParser(description="""Stepper simulator. Given a stepper
		spreadsheet and reticle simulates the expected wafer layout generated by
		the stepper.""")
parser.add_argument("--reticle-list", metavar='reticle_list', type=str,\
	nargs='+', help='list of reticle cells to be referenced')
parser.add_argument("--spreadsheet", metavar='fn_spreadsheet', type=str,
	help='name of the stepper spreadsheet file')
parser.add_argument("--mask-file", metavar='fn_mask', type=str,\
	help="""name of the file in GDS format with the reticles""")
parser.add_argument("--out-file", metavar='fn_out', nargs='?', type=str,\
		help="""name of the file in which to store the stepper sim output.
		By default appends '_sim-output' to the mask-file name.""")

#print("Loading mask...")
#lib = gdspy.GdsLibrary()
#lib.read_gds(fn_mask)
#pmask = lib.cell_dict[mask_cell].flatten().elements
#c = gdspy.Cell(mask_cell)
#c.add(pmask)
#
#print("Loading spreadsheet...")
#wb = openpyxl.load_workbook(fn_spreadsheet, data_only=True)
#ws = wb[wb.sheetnames[0]]
#
## The first row with information
#r = 15
#
#next_layer_num = 1
#layer_nums = {}

def arg_check(f):
	"""Checks for None values and returns 0 by default"""
	def type_check(val):
		try:
			r = f(val)
		except TypeError:
			r = 0
		return r
	return type_check

@arg_check
def to_int(val):
	return int(val)

@arg_check
def to_float(val):
	return float(val)

def to_list(val):
		if not val:
				return []
		return [int(x) for x in str(val).split(',')]

def get_shift_info(ws):
	startrow = 3
	row = startrow
	while True:
		name = ws['A' + str(row)].value
		try:
			assert name != None
		except AssertionError:
			return
		ncols = to_int(ws['C' + str(row)].value)
		nrows = to_int(ws['D' + str(row)].value)
		col_shifts = [0]*ncols
		row_shifts = [0]*nrows

		for i in range(ncols):
			col_shifts[i] = 1e3*to_float(ws['M' + str(row+i)].value)
		for j in range(nrows):
			row_shifts[j] = 1e3*to_float(ws['R' + str(row+j)].value)

		row += (max([ncols, nrows]) + 2)
		yield col_shifts, row_shifts





def run_sim(fn_spreadsheet, mask_cells):
	print("Loading spreadsheet...")
	wb = openpyxl.load_workbook(fn_spreadsheet, data_only=True)
	ws = wb[wb.sheetnames[0]]
	shift_ws = wb[wb.sheetnames[1]]
	shift_tracker = get_shift_info(shift_ws)
	# The first row with information
	r = 15

	next_layer_num = 1
	layer_cells = {}
	layer_nums = {}


	while True:

		layer = ws['A' + str(r)].value

		# If we find an empty line, we are done
		if not layer:
				break
		name_ = ws['C' + str(r)].value
		name = "row" + str(r) + '_' + name_
		print("Working on cell " + name)

		# Assign a number to each layer name
		if layer not in layer_nums:
			layer_nums[layer] = next_layer_num
			layer_cells[layer] = gdspy.Cell("layer_" + layer.lower())
			next_layer_num += 1
		if layer not in layer_cells:
			layer_cells[layer] = gdspy.Cell("layer_" + layer.lower())

		# Assuming the column layout from Albert's generated code
		b_xl = 1e3*float(ws['H' + str(r)].value)
		b_xr = 1e3*float(ws['I' + str(r)].value)
		b_yu = 1e3*float(ws['J' + str(r)].value)
		b_yd = 1e3*float(ws['K' + str(r)].value)
		ws_x = 1e3*float(ws['O' + str(r)].value)
		ws_y = 1e3*float(ws['P' + str(r)].value)
		ar_nx = to_int(ws['T' + str(r)].value)
		ar_ny = to_int(ws['U' + str(r)].value)
		ar_cx = 1e3*to_float(ws['V' + str(r)].value)
		ar_cy = 1e3*to_float(ws['W' + str(r)].value)
		ar_xp = 1e3*to_float(ws['X' + str(r)].value)
		ar_yp = 1e3*to_float(ws['Y' + str(r)].value)
		reticle_name = ws['S' + str(r)].value
		skip_c = to_list(ws['Z' + str(r)].value)
		skip_r = to_list(ws['AA' + str(r)].value)
		to_shift = bool(ws['AB' + str(r)].value)
		if to_shift:
			col_shifts, row_shifts = next(shift_tracker)
		# Bladed region
		pblades = gdspy.Rectangle((b_xl,b_yu),(b_xr,b_yd))

		c = gdspy.Cell("blades_" + name)
		c.add(gdspy.CellReference(reticle_name))
		c.add(pblades)

		try:
			pcell = gdspy.fast_boolean(mask_cells[reticle_name], pblades, 'and',\
				layer=layer_nums[layer])
		except KeyError:
			print ("failed for cell " + name)

		# Make sure we are left with geometry after blading
		assert pcell is not None

		c = gdspy.Cell(name)
		c.add(pcell)

		c = gdspy.Cell("patch_" + name)

		# Adjust the center.  The stepper centers arrays, gdspy defines the corner.
		x0 = - ws_x - (ar_nx - 1) * ar_xp / 2.0
		y0 = - ws_y + (ar_ny - 1) * ar_yp / 2.0

		for x in range(ar_nx):
			dx, dy = 0, 0
			if to_shift:
				dx = col_shifts[x]
			if (x+1) in skip_c:
				continue
			for y in range(ar_ny):
				if to_shift:
					dy = row_shifts[y]
				if (y+1) in skip_r:
					continue
				c.add(gdspy.CellReference(name, (x0 + x*ar_xp + dx,\
					y0 - y*ar_yp + dy)))

		layer_cells[layer].add(gdspy.CellReference("patch_" + name, (ar_cx, ar_cy)))

		r += 1

	cwafer = gdspy.Cell('wafer')
	for k, v in layer_cells.items():
			print(v.name)
			cwafer.add(gdspy.CellReference(v.name))

	# Wafer border
	r1 = gdspy.Round((0,0),radius=50.1*1000, layer = 0, number_of_points=50)
	r2 = gdspy.Round((0,0),radius=49.9*1000, layer = 0, number_of_points=50)
	r = gdspy.fast_boolean(r1, r2, 'not', layer=0)
	cwafer.add(r)

	print("Writing file...")
	gdspy.write_gds(fn_out, unit=1.0e-6, precision=1.0e-9)
	# gdspy.LayoutViewer()
	print("Done!")

def get_size(cell):
	(xmin, ymin), (xmax, ymax) = cell.get_bounding_box()
	return (xmax - xmin), (ymax - ymin)


# Makes a rectangle on a default layer
def make_rectangle(width,length):
	return gdspy.Rectangle([-width/2, length/2], [width/2, -length/2], layer=0)

def fill_empty_space(cell, width, length):
	filler = make_rectangle(width, length)
	subcells = cell.elements
	for subcell in subcells:
		dx, dy = get_size(subcell)
		subrect = make_rectangle(dx, dy)
		subrect.translate(*subcell.origin)
		filler = gdspy.fast_boolean(filler, subrect, 'xor',layer=0)
		# print (subcell)
	return filler

def load_mask(fn_mask, rlist):
	mask_cells = {}
	lib = gdspy.GdsLibrary()
	lib.read_gds(fn_mask)
	for rname in rlist:
		c = gdspy.Cell(rname)
		mask_elems = lib.cell_dict[rname].elements
		mask_elems = [x for x in mask_elems if type(x) in allowed_element_types]
		c.add(mask_elems)
		filler = fill_empty_space(c, mask_width, mask_length)
		c.add(filler)
		pmask = c.flatten().elements
		main_lib.add(c)
		mask_cells[rname] = pmask
	return mask_cells

if __name__=='__main__':
	args = parser.parse_args()
	print (args)
	fn_spreadsheet = args.spreadsheet
	fn_mask = args.mask_file
	reticle_list = args.reticle_list
	fn_out = args.out_file if args.out_file else os.path.splitext(fn_mask)[0] + '_sim-output.gds'
	print("Loading mask...")
	mask_cells = load_mask(fn_mask, reticle_list)
	run_sim(fn_spreadsheet, mask_cells)
