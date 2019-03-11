import gdspy
import numpy as np
import pdb
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from string import ascii_uppercase

allowed_element_types = set([gdspy.Cell, gdspy.CellReference, gdspy.CellArray])
default = np.zeros(2)
scale = 1000

class cellNode():

	"""
	Keeps track of a single cell and its origin relative to the root
	of the cell hierarchy tree
	"""
	def __init__(self, cell, origin=default):
		self.cell = cell
		self.origin = origin

	def update_origin(self, new_origin):
		self.origin += new_origin

# scales all the values in an array by the scale
def scalearr(arr, scale):
  return np.asarray(arr)/scale


# Need this functionality to work with excel spreadsheets
chars = list(ascii_uppercase)
chars2 = [x + y for x in chars for y in chars]
chars += chars2
char2num = {char:(ind+1) for ind, char in enumerate(chars)}
num2char = {(ind+1):char for ind, char in enumerate(chars)}

al = Alignment(horizontal="center", wrap_text=False)
font = Font(name='Times New Roman', size=10)
thin = Side(border_style='thin', color='000000')
border = Border(top=thin, left=thin, right=thin, bottom=thin)
fill = PatternFill(fill_type=None, start_color='FFFFFFFF', end_color='FF000000')

#def as_text(value): return str(value) if value is not None else ""

def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
	"""
	:param ws:  Excel worksheet instance
	:param range: An excel range to style (e.g. A1:F20)
	:param border: An openpyxl Border
	:param fill: An openpyxl PatternFill or GradientFill
	:param font: An openpyxl Font object
	"""



	rows = ws[cell_range]
	dims = {}
	for row in rows:
		for cell in row:
			cell.font = font
			cell.border = border
			cell.fill = fill
			cell.alignment = alignment

def populate_column(sheet, col, startrow, dataset):
	N = len(dataset)
	for index in range(N):
		row = startrow + index
		sheet.cell(column = col, row=row, value = dataset[index])


class PatchTable():
    """
    Class representing a collection of the jobs to be shot by the stepper
    programmer. The PatchTable is used to generate a spreadsheet that is then
    entered into the stepper.

    Attributes:
    -------------
        tvpa_shift,
        aga_x_shift,
        aga_y_shift: The relative positions of the TVPA and AGA marks with
            respect to the center of the alignment marks cell. Assumes that the
            alignment marks cell always starts with "alignment". Customize this
            before generating the spreadsheet.

            By default:
                tvpa_shift = np.array([0.125, 0])
                aga_x_shift = np.array([-0.125, 0])
                aga_y_shift = np.array([-0.225, 0])

        align_layer:
            First layer to be shot. This is the layer on which the
            alignment cell is to be placed on.
        shots: list of Shot objects, each representing a job on the stepper.
        filename: name of the workbook in which to save the spreadsheet.
        wb: Excel workbook in which the spreadsheet is saved.
        ws: active Excel worksheet instance in which to save the data.
        shift_ws: secondary worksheet that holds all the array shift information.
        layers: array of all the layer numbers of the stepper jobs.
        names: names of all the stepper jobs.
        mask_names: array of the names of all the reticle cells used to source the
            jobs.
        mask_shifts:
        cell_sizes:
        cell_shifts:
        cell_bboxs:
        array_sizes:
        array_centers:
        array_stepsizes:
        array_shifted:
        array_positions:
        calculated_xys:
        desired_xys:
        shifts_xys:
        num_shots:
        xl:
        xr:
        yu:
        yd:
        wafer_shifts:


    Methods:
    ------------
        generate_spreadsheet(): Creates and saves a spreadsheet under the
            name self.filename.

    Example:
    ----------
        patchtable = PatchTable(shotlist, "stepperprog.xlsx") where the shotlist has
            been generated using gen_patches_table(...).
        patchtable.generate_spreadsheet()
    """
	# Some class constants that define the relative positions of the TVPA and
	# AGA Marks with respect to the center of the alignment marks cell
	# These are defined for alignment_marks_patch_new_r
	tvpa_shift = np.array([0.125, 0])
	aga_x_shift = np.array([-0.125, 0])
	aga_y_shift = np.array([-0.225, 0])

	# For alignment_marks_patch_r
	#tvpa_shift = np.array([0.16875, 0])
	#aga_x_shift = np.array([-0.08125, 0])
	#aga_y_shift = np.array([-0.18125, 0])


	def __init__(self, shotlist, wb_filename="stepper.xlsx"):
		"""
        PatchTable(shotlist, wb_filename)
        ----------------------------------------------------------

        Creates a PatchTable class initialized with a list of Shot() objects and
        the name of the workbook filename.

        Params:
            shotlist (list):  List of Shot() objects to be represented on the
                spreadsheet.
            wb_filename (str): Name of the excel file in which the stepper
                programming information is to be saved.

        """
        self.shots = shotlist
		self.filename = wb_filename
		self.wb = Workbook()
		self.layers = np.array(list(map(lambda x: x.get_layer(),\
				self.shots)))
		self.names = np.array(list(map(lambda x: x.get_name(),\
				self.shots)))
		self.mask_shifts = np.array(list(map(lambda x: x.get_mask_shift(),\
				self.shots)))
		self.cell_sizes = np.array(list(map(lambda x: x.get_cell_size(),\
				self.shots)))
		self.cell_shifts = np.array(list(map(lambda x: x.get_cell_shift(),\
				self.shots)))
		self.cell_bboxs = np.array(list(map(lambda x: x.get_cell_bbox(),\
				self.shots)))
		self.mask_names = np.array(list(map(lambda x: x.get_maskname(),\
				self.shots)))
		self.array_sizes = np.array(list(map(lambda x: x.get_array_size(),\
				self.shots)))
		self.array_centers = np.array(list(map(lambda x:\
				x.get_array_center(), self.shots)))
		self.array_stepsizes = np.array(list(map(lambda x:\
				x.get_array_stepsize(), self.shots)))
		self.array_shifted = np.array(list(map(lambda x:\
				x.get_array_shifts(), self.shots)))
		self.array_positions = np.array(list(map(lambda x:\
				x.get_array_positions(), self.shots)))
		self.calculated_xys = np.array(list(map(lambda x:\
				x.get_calculated_xy(), self.shots)))
		self.desired_xys = np.array(list(map(lambda x:\
				x.get_desired_xy(), self.shots)))
		self.shifts_xys = np.array(list(map(lambda x:\
				x.get_shift_xy(), self.shots)))
		self.all_dropped_rows = np.array(list(map(lambda x:\
				x.get_dropped_rows(), self.shots)))
		self.all_dropped_columns = np.array(list(map(lambda x:\
				x.get_dropped_columns(), self.shots)))
		self.num_shots = len(self.layers)
		self.xl = self.mask_shifts[:, 0] + self.cell_bboxs[:,0,0]
		self.xr = self.mask_shifts[:, 0] + self.cell_bboxs[:,1,0]
		self.yu = self.mask_shifts[:, 1] + self.cell_bboxs[:,1,1]
		self.yd = self.mask_shifts[:, 1] + self.cell_bboxs[:,0,1]
		self.align_layer = self.layers[0]

		self.wafer_shifts = self.mask_shifts - self.cell_shifts
		self.ws = self.wb.active
		self.ws.title = 'Patches' #get the active worksheet in which to save the data
		self.shift_ws = self.wb.create_sheet('Shift_Calculator')

	def populate_alignment_info(self):
		self.ws['B4'] = 'AGA Marks'
		self.ws['H4'] = 'TVPA'
		self.ws['B5'] = 'direction'
		self.ws['C5'] = 'layer'
		self.ws['D5'] = 'x'
		self.ws['E5'] = 'y'
		self.ws['G5'] = 'layer'
		self.ws['H5'] = 'x'
		self.ws['I5'] = 'y'
		self.ws['B6'] = 'y'
		self.ws['B7'] = 'x'
		self.ws['C6'] = self.align_layer
		self.ws['C7'] = self.align_layer
		self.ws['G6'] = self.align_layer

		# Get the relative position of the AGA and TVPA marks from the cell
		# shift of the alignment layer cell
		align_shift = self.cell_shifts[0]
		tvpa_pos = align_shift + PatchTable.tvpa_shift
		aga_x_pos = align_shift + PatchTable.aga_x_shift
		aga_y_pos = align_shift + PatchTable.aga_y_shift

		self.ws['D6'] = aga_y_pos[0]
		self.ws['E6'] = aga_y_pos[1]
		self.ws['D7'] = aga_x_pos[0]
		self.ws['E7'] = aga_x_pos[1]
		self.ws['H6'] = tvpa_pos[0]
		self.ws['I6'] = tvpa_pos[1]

	def populate_shiftcalc(self):
		ws = self.shift_ws

		cell_range = "A1:AA500"
		style_range(ws, cell_range, border=border, fill=fill,\
				font=font, alignment=al)
		ws['A2'] = 'Job'
		ws['B2'] = 'patch'
		ws['C1'] = 'Array'
		ws['E1'] = 'Centered (mm)'
		ws['G1'] = 'Pitch (mm)'
		ws['J1'] = 'Column'
		ws['J2'] = 'position'
		ws['O1'] = 'Row'
		ws['O2'] = 'position'
		ws['K1'] = 'Calculate'
		ws['P1'] = 'Calculate'
		ws['L1'] = 'Desired'
		ws['Q1'] = 'Desired'
		ws['M1'] = 'Shift in'
		ws['R1'] = 'Shift in'
		ws['C2'] = 'C'
		ws['D2'] = 'R'
		ws['E2'] = 'x'
		ws['G2'] = 'x'
		ws['K2'] = 'x'
		ws['L2'] = 'x'
		ws['M2'] = 'x'
		ws['F2'] = 'y'
		ws['H2'] = 'y'
		ws['P2'] = 'y'
		ws['Q2'] = 'y'
		ws['R2'] = 'y'
		N = len(self.array_shifted)
		start_row = 3
		row = start_row
		for i in range(N):
			if not self.array_shifted[i]:
				continue
			ws['A' + str(row)] = self.names[i]
			ws['B' + str(row)] = self.layers[i]
			ws['C' + str(row)] = self.array_sizes[:,0][i]
			ws['D' + str(row)] = self.array_sizes[:,1][i]
			ws['E' + str(row)] = self.array_centers[:, 0][i]
			ws['F' + str(row)] = self.array_centers[:, 1][i]
			ws['G' + str(row)] = self.array_stepsizes[:, 0][i]
			ws['H' + str(row)] = self.array_stepsizes[:, 1][i]

			Ncols = len(self.array_positions[:, 0][i])
			Nrows = len(self.array_positions[:, 1][i])
			#print (self.calculated_xys[:, 0][i])
			#print (self.calculated_xys[:, 1][i])
			crow = row
			for j in range(Ncols):
				ws['J' + str(crow)] = self.array_positions[:, 0][i][j]
				ws['K' + str(crow)] = self.calculated_xys[:, 0][i][j]
				ws['L' + str(crow)] = self.desired_xys[:, 0][i][j]
				ws['M' + str(crow)] = self.shifts_xys[:, 0][i][j]
				crow += 1
			rrow = row
			for j in range(Nrows):
				ws['O' + str(rrow)] = self.array_positions[:, 1][i][j]
				ws['P' + str(rrow)] = self.calculated_xys[:, 1][i][j]
				ws['Q' + str(rrow)] = self.desired_xys[:, 1][i][j]
				ws['R' + str(rrow)] = self.shifts_xys[:, 1][i][j]
				rrow += 1
			row += max([Ncols, Nrows]) + 2
			#row += 1
		ws.merge_cells('C1:D1')
		ws.merge_cells('E1:F1')
		ws.merge_cells('G1:H1')
		ws.merge_cells('E1:F1')

		# Adjust the spreadsheet to ensure that the width of the cells is enough
		for column_cells in ws.columns:
			length = max(len(str(cell.value or ""))+2 for cell in column_cells)
			if length < 6: length = 6
			ws.column_dimensions[column_cells[0].column].width = length



	def generate_spreadsheet(self):
		# Write down the TVPA and AGA information at the top of the spreadsheet
		self.populate_alignment_info()

		# Format the spreadsheet before you write the data. Formatting multiple
		# cells seems to overwrite the data
		self.startrow = 15
		self.endrow = 15 + self.num_shots

		cell_range = "A4:AB%d"%(self.endrow)
		style_range(self.ws, cell_range, border=border, fill=fill,\
				font=font, alignment=al)

		populate_column(self.ws, char2num['A'], 15, self.layers)
		populate_column(self.ws, char2num['C'], 15, self.names)
		populate_column(self.ws, char2num['D'], 15, self.mask_shifts[:,0])
		populate_column(self.ws, char2num['E'], 15, self.mask_shifts[:,1])
		populate_column(self.ws, char2num['F'], 15, self.cell_sizes[:, 0])
		populate_column(self.ws, char2num['G'], 15, self.cell_sizes[:, 1])
		populate_column(self.ws, char2num['M'], 15, self.cell_shifts[:, 0])
		populate_column(self.ws, char2num['N'], 15, self.cell_shifts[:, 1])

		populate_column(self.ws, char2num['H'], 15, self.xl)
		populate_column(self.ws, char2num['I'], 15, self.xr)
		populate_column(self.ws, char2num['J'], 15, self.yu)
		populate_column(self.ws, char2num['K'], 15, self.yd)

		populate_column(self.ws, char2num['O'], 15, self.wafer_shifts[:,0])
		populate_column(self.ws, char2num['P'], 15, self.wafer_shifts[:,1])
		populate_column(self.ws, char2num['S'], 15, self.mask_names)

		populate_column(self.ws, char2num['T'], 15, self.array_sizes[:,0])
		populate_column(self.ws, char2num['U'], 15, self.array_sizes[:,1])
		populate_column(self.ws, char2num['V'], 15, self.array_centers[:, 0])
		populate_column(self.ws, char2num['W'], 15, self.array_centers[:, 1])
		populate_column(self.ws, char2num['X'], 15, self.array_stepsizes[:, 0])
		populate_column(self.ws, char2num['Y'], 15, self.array_stepsizes[:, 1])
		populate_column(self.ws, char2num['Z'], 15, self.all_dropped_columns)
		populate_column(self.ws, char2num['AA'], 15, self.all_dropped_rows)
		populate_column(self.ws, char2num['AB'], 15, self.array_shifted)

		# Adjust the spreadsheet to ensure that the width of the cells is enough
		for column_cells in self.ws.columns:
			length = max(len(str(cell.value or ""))+2 for cell in column_cells)
			if length < 6: length = 6
			self.ws.column_dimensions[column_cells[0].column].width = length

		self.ws['A14'] = 'Layer'
		self.ws['B14'] = 'Description'
		self.ws['C14'] = 'Cell name'
		self.ws['D13'] = 'Mask Shift'
		self.ws['D14'] = 'x'
		self.ws['E14'] = 'y'
		self.ws['F13'] = 'Cell size'
		self.ws['F14'] = 'x'
		self.ws['G14'] = 'y'
		self.ws['M13'] = 'Cell Shift'
		self.ws['M14'] = 'x'
		self.ws['N14'] = 'y'
		self.ws['H13'] = 'Blade Coordinates'
		self.ws['H14'] = 'xl'
		self.ws['I14'] = 'xr'
		self.ws['J14'] = 'yu'
		self.ws['K14'] = 'yd'
		self.ws['O13'] = 'Wafer Shift'
		self.ws['O14'] = 'x'
		self.ws['P14'] = 'y'
		self.ws['Q13'] = 'Patch'
		self.ws['Q14'] = 'Frontside'
		self.ws['R14'] = 'Dose'
		self.ws['S14'] = 'Mask'
		self.ws['T12'] = 'Alternate Array Layout'
		self.ws['T13'] = 'size'
		self.ws['T14'] = 'C'
		self.ws['U14'] = 'R'
		self.ws['V13'] = 'center'
		self.ws['V14'] = 'x'
		self.ws['W14'] = 'y'
		self.ws['X13'] = 'step size'
		self.ws['X14'] = 'x'
		self.ws['Y14'] = 'y'
		self.ws['Z12'] = 'To Skip'
		self.ws['Z13'] = '(1-indexed, from UL)'
		self.ws['Z14'] = 'C'
		self.ws['AA14'] = 'R'

		self.ws.merge_cells('D13:E13')
		self.ws.merge_cells('F13:G13')
		self.ws.merge_cells('M13:N13')
		self.ws.merge_cells('H13:K13')
		self.ws.merge_cells('O13:P13')
		self.ws.merge_cells('T12:Y12')
		self.ws.merge_cells('T13:U13')
		self.ws.merge_cells('V13:W13')
		self.ws.merge_cells('X13:Y13')
		self.ws.merge_cells('Z12:AA12')
		self.ws.merge_cells('Z13:AA13')

		self.populate_shiftcalc()

		self.wb.save(self.filename)

class Shot():
	"""
    Class represents a single shot that is to be made by the stepper. Holds all
    the information present in a single row of the stepper programming
    spreadsheet.

    -------------------------------------------------------------------------

    Attributes:
    ------------
        def_layers: Dictionary mapping between layer names and layer numbers.
        inv_layers = Same dictionary structure as def_layers but with the layer
            numbers as the keys and the layer names as the values.
        ordering: list of the order in which the layers are to be shot. Used to
            define an ordering for sorting shots.
        cell: gdspy cell that represents the shot.
        cellname: name of the gdspy cell.
        layer: Layer number of the cell. Note that a shot MUST have only one layer.
               If cell has more than 1 layer, a RuntimeError is raised.
        cell_bbox: 2x2 array that is the bounding box of the cell
        cell_size: array representing the size of the cell
        mask_name: name of the reticle cell from which the shot is to be made
        cell_shift: position of the cell in the global overlay to be made
        isArray: whether the shot is to be made into an array or not.
        maskcell:
        maskcellsize:
        mask_shift:
        mask_cellbbox:

        If the cell is arrayed, checks the kwargs:
        kwargs:
        ---------
            ncols:
            nrows:
            center:
            xspacing:
            yspacing:
            is_shifted:
            If is_shifted is True,
                column_pos:
                row_pos:
                calculated_x:
                calculated_y:
                desired_x:
                desired_y:
                shift_x:
                shift_y:

    """
    def_layers = {"Thin Gold":1, "PRO1":2, "ALUMINUM":3, "LSNSUB":4, "LSN1":5,\
			"120nm_NbWiring":6, "STEPPER":7, "400nm_NbWiring":8, "ILD":9,\
			"XeF2":10, "GP":12, 'Wafer Outline':22}
	inv_layers = {value:key for key, value in def_layers.items()}
	layer_list = sorted(list(def_layers.values()))
	index = list(range(len(layer_list)))
	ordering = {x:y for x, y in zip(layer_list, index)}

	def update_layers(layers_dict):
		Shot.def_layers = layers_dict
		Shot.inv_layers = {value:key for key, value in Shot.def_layers.items()}
		Shot.layer_list = sorted(list(Shot.def_layers.values()))
		Shot.index = list(range(len(Shot.layer_list)))
		Shot.ordering = {x:y for x, y in zip(Shot.layer_list, Shot.index)}

	def update_layerorder(layer_order):
		Shot.index = list(range(len(layer_order)))
		Shot.ordering = {x:y for x, y in zip(layer_order, Shot.index)}


	def __init__(self, cell, cell_shift, cell_size, cell_bbox, mask_name="", isArray=False, **kwargs):
		try:
			assert(type(cell) == gdspy.Cell or type(cell) == gdspy.CellReference)
		except AssertionError:
			print ("Please provide a valid cell to construct the shot")
		self.cell = cell
		self.cellname = cell.name
		layers = list(cell.get_layers())
		if len(layers) > 1 :
		  raise RuntimeError ("Cannot construct a shot from a cell with multiple layers")
		self.layer = layers[0]
		self.cell_bbox = np.asarray(cell_bbox)
		self.cell_size = np.asarray(cell_size)
		self.mask_name = mask_name
		self.cell_shift = np.asarray(cell_shift)
		self.isArray = isArray
		#self.blade_coords = {'xl':, 'xr':, 'yt':, 'yb':}
		if isArray:
			self.ncols = kwargs['num_cols']
			self.nrows = kwargs['num_rows']
			self.center = np.asarray(kwargs['center'])
			self.xspacing = kwargs['xspacing']
			self.yspacing = kwargs['yspacing']
			self.is_shifted = False
			self.dropped_rows = np.empty((0,), dtype=np.int64)
			self.dropped_cols = np.empty((0,), dtype=np.int64)
			try:
				if kwargs['is_shifted']:
					self.is_shifted = True
					self.column_pos = kwargs['column_pos']
					self.row_pos = kwargs['row_pos']
					self.calculated_x = kwargs['calculated_x']
					self.calculated_y = kwargs['calculated_y']
					self.desired_x = kwargs['desired_x']
					self.desired_y = kwargs['desired_y']
					self.shift_x = kwargs['shift_x']
					self.shift_y = kwargs['shift_y']
				elif kwargs['dropRC_args']['to_drop']:
					self.center = np.asarray(kwargs['dropRC_args']['center'])
					self.xspacing = kwargs['dropRC_args']['xspacing']
					self.yspacing = kwargs['dropRC_args']['yspacing']
					self.ncols = kwargs['dropRC_args']['num_cols']
					self.nrows = kwargs['dropRC_args']['num_rows']
					self.dropped_rows = kwargs['dropRC_args']['dropped_rows']
					self.dropped_cols = kwargs['dropRC_args']['dropped_cols']

			except KeyError:
				pass
				#do nothing
				#print ('Nope. Nothing here')
		# defaults
		self.maskcell=""
		self.maskcellsize=default
		self.mask_shift=default
		self.mask_cell_bbox = [[0.0,0.0],[0.0,0.0]]

	def update_mask_location(self, maskcellref, maskname):
		self.mask_name = maskname
		maskcellname = maskcellref.ref_cell.name
		assert(maskcellname.startswith(self.cellname))

		self.maskcell = maskcellref.ref_cell
		self.maskcellsize = scalearr(get_size(maskcellref), scale)
		self.mask_cell_bbox = scalearr(self.maskcell.get_bounding_box(), scale)
		self.mask_shift = scalearr(maskcellref.origin, scale)

	def __repr__(self):
		if not self.isArray:
			return "Shot({0}, {1}, {2}, {3})".format(self.cellname,\
					self.cell_shift,self.mask_shift, self.maskcellsize)
		return "Shot({0}, {1}, {2}, {3}, {4}, {5})".format(self.cellname,\
			self.center, self.mask_shift, self.maskcellsize,\
			(self.ncols, self.nrows),(self.xspacing, self.yspacing))

	def __eq__(self, other):
		if Shot.ordering[self.layer] == Shot.ordering[other.layer]:
			return self.cellname == other.cellname
		return False

	def __lt__(self, other):
		if Shot.ordering[self.layer] == Shot.ordering[other.layer]:
			if self.cellname.startswith('alignment'): return True
			if other.cellname.startswith('alignment'): return False
			return self.cellname < other.cellname
		return Shot.ordering[self.layer] < Shot.ordering[other.layer]

	def get_layer(self):
		return Shot.inv_layers[self.layer]

	def get_name(self):
		return self.cellname

	def get_maskname(self):
		return self.mask_name

	def get_mask_shift(self):
		return self.mask_shift

	def get_cell_size(self):
		return self.maskcellsize

	def get_cell_shift(self):
		return self.cell_shift

	def get_cell_bbox(self):
		return self.mask_cell_bbox

	def get_array_size(self):
		if not self.isArray: return None, None
		return self.ncols, self.nrows

	def get_array_center(self):
		if not self.isArray: return None, None
		return self.center

	def get_array_stepsize(self):
		if not self.isArray: return None, None
		return self.xspacing, self.yspacing

	def get_array_shifts(self):
		if not self.is_shifted:
			return ""
		else:
			return "shift"

	def get_array_positions(self):
		if not self.is_shifted: return None, None
		return self.column_pos, self.row_pos

	def get_calculated_xy(self):
		if not self.is_shifted: return None, None
		return self.calculated_x, self.calculated_y

	def get_desired_xy(self):
		if not self.is_shifted: return None, None
		return self.desired_x, self.desired_y

	def get_shift_xy(self):
		if not self.is_shifted: return None, None
		return self.shift_x, self.shift_y

	def get_dropped_columns(self):
		if self.dropped_cols.size == 0: return None
		#return list(self.dropped_cols)
		return ','.join([str(x) for x in self.dropped_cols])

	def get_dropped_rows(self):
		if self.dropped_rows.size == 0: return None
		#return list(self.dropped_rows)
		return ','.join([str(x) for x in self.dropped_rows])

def translate(shot, shift):
	shot.cell_shift -= np.asarray(shift)

def get_size(cell):
	(xmin, ymin), (xmax, ymax) = cell.get_bounding_box()
	dx = (xmax - xmin)
	dy = (ymax - ymin)
	return np.array([dx, dy])

def get_center(cell):
	(xmin, ymin), (xmax, ymax) = cell.get_bounding_box()
	x0 = (xmax + xmin)//2
	y0 = (ymax + ymin)//2
	return np.array([x0, y0])

def inv_mask_cellname(shot, ending):
	if shot.cellname.endswith('_r'):
		name = shot.cellname
	elif shot.cellname.endswith('inv'):
		name = shot.cellname
	elif shot.cellname.endswith('-r'):
		name = shot.cellname
	else:
		name = shot.cellname + ending
	return name

def find_match(ref_cell, match_name):
	try:
		match = list(filter(lambda x: x.ref_cell.name == match_name,
			ref_cell))[0]
		return match
	except IndexError:
		return None

def gen_patches_table(globaloverlay, mask_list, ignored_cells, layer_dict=None,\
		layer_order=None, cellsInverted=True, invcell_ending='_inv'):
    """
    Function for generating a list of shots using the layout cell and a list of
    reticle cells. Descends down the hierarchy of cells in the layout cell and
    identifies the positions, sizes and array properties needed to generate the
    full layout. Each unique cell on the layout MUST have a matching cell in one
    of the reticle cells on the reticle list.

    Args:
    ---------------
        globaloverlay (gdspy.Cell): layout cell that is to be fabricated.
        mask_list ([gdspy.Cell, ...]): list of all the reticle_cells.
        ignored_cells ([gdspy.Cell, ...]): list of all subcells of the layout
            cell that shouldn't be shot.
        layer_dict (dict): dictionary mapping layer names: layer numbers.
        layer_order (list): list of all the layer numbers in the order in which they
            are to be shot.
        cellsInverted (bool): Boolean value representing whether the cells on the
            reticles are the inverted versions of the cells on the layout.
        invcell_ending (str): Naming convention used to distinguish cells on the
            layout and the inverted cells on the reticle.
    Returns:
    ------------------
        (list): A list of all the shots for the stepper program.

    """
	Shot.update_layers(layer_dict)
	if layer_order:
		Shot.update_layerorder(layer_order)
	gcomponents = globaloverlay.elements
	allshots = []
	#pdb.set_trace()
	for component in gcomponents:
		if type(component) not in allowed_element_types: continue
		if component.ref_cell.name in ignored_cells: continue
		shots_made = makeshot(component, mask_list=mask_list,\
			ignored_cells=ignored_cells)
		#print (component.ref_cell.name, len(shots_made))
		allshots.extend(shots_made)

	mcomponents = {}
	for mask in mask_list:
		mcomponents[mask.name] = [x for x in mask.elements if type(x) in allowed_element_types]

	for shot in allshots:
		inv_name = inv_mask_cellname(shot, invcell_ending)
		name = shot.cellname
		#print (inv_name, name)
		for mask in mcomponents:
			inv_match = find_match(mcomponents[mask], inv_name)
			match = find_match(mcomponents[mask], name)
			#print (inv_match, match)
			if cellsInverted and inv_match:
				shot.update_mask_location(inv_match, mask)
			elif cellsInverted and match:
				shot.update_mask_location(match, mask)
			else:
				continue

	allshots.sort()
	return allshots

empty_dict = dict()

def get_cell_asymmetry(cell):
	(xmin, ymin), (xmax, ymax) = cell.get_bounding_box()
	dx = xmax - np.abs(xmin)
	dy = ymax - np.abs(ymin)
	return [dx, dy]

def check_dropped_rows_columns(array_x, array_y, center):
	dropped_rc_args = {}
	dropped_rc_args['to_drop'] = True
	# I'll assume all dimensions have at most 3dp
	diff_x = np.diff(array_x*1000).astype(int)
	diff_y = np.diff(array_y*1000).astype(int)
	dx, dy = (0,0)
	drop_rows = False
	drop_cols = False
	if np.all(diff_x) != 0:
		gcd_x = np.gcd.reduce(diff_x)
		dx = gcd_x/1000
		cols = np.hstack(([0], np.cumsum(diff_x//gcd_x)))
		ncols = cols[-1]+1
		arr_x = np.zeros(ncols, dtype=bool)
		arr_x[cols] = True
		if not np.all(arr_x): drop_cols = True
	else:
		arr_x = np.ones_like(array_x, dtype=bool)
		ncols = 1
	if np.all(diff_y) != 0:
		gcd_y = np.gcd.reduce(diff_y)
		dy = gcd_y/1000
		# Need -diff_y because as the index increases the value of y decreases
		rows = np.hstack(([0], np.cumsum(-diff_y//gcd_y)))
		nrows = rows[-1]+1
		arr_y = np.zeros(nrows, dtype=bool)
		arr_y[rows] = True
		if not np.all(arr_y): drop_rows = True
	else:
		nrows = 1
		arr_y = np.ones_like(array_y, dtype=bool)
	if ncols > 20 or nrows > 20:
		dropped_rc_args['to_drop'] = False
	x0 = np.around((array_x[-1] + array_x[0])/2, 3)
	y0 = np.around((array_y[-1] + array_y[0])/2, 3)
	if center[0] == x0: x0 = 0
	if center[1] == y0: y0 = 0
	dropped_rc_args['center'] = (x0, y0)
	dropped_rc_args['xspacing'] = dx
	dropped_rc_args['yspacing'] = dy
	dropped_rc_args['num_cols'] = ncols
	dropped_rc_args['num_rows'] = nrows
	dropped_rc_args['columns'] =  arr_x
	dropped_rc_args['rows'] = arr_y
	dropped_rc_args['to_drop_rows'] = drop_rows
	dropped_rc_args['to_drop_cols'] = drop_cols
	dropped_rc_args['dropped_rows'] = np.argwhere(np.logical_not(arr_y)).flatten()+1
	dropped_rc_args['dropped_cols'] = np.argwhere(np.logical_not(arr_x)).flatten()+1
	return dropped_rc_args


def get_array_shifts(element, parent_args):
	# Unpack the parent array info for easier computation
	Nc,Nr = parent_args['num_cols'], parent_args['num_rows']
	X0, Y0 = parent_args['center']
	DX, DY = parent_args['xspacing'], parent_args['yspacing']
	ptx, pty = parent_args['to_shift']
	# Get info of the child array
	nc, nr = element.columns, element.rows
	x0, y0 = scalearr(get_center(element) - get_center(element.ref_cell), scale)
	dx, dy = scalearr(element.spacing, scale)
	mycenter = np.array([x0, y0])
	# Compute the new array properties
	Mc, Mr = Nc*nc, Nr*nr
	Mx = np.arange(Mc) + 1
	My = np.arange(Mr) + 1
	nx = (Mx - 1)%nc + 1
	Nx = ((Mx - nx)/nc + 1).astype(int)
	ny = (My - 1)%nr + 1
	Ny = ((My - ny)/nr + 1).astype(int)
	if parent_args['is_shifted']:
		pshiftx = parent_args['shift_x'][Nx - 1]
		pshifty = parent_args['shift_y'][Ny - 1]

	xspacing = DX//nc
	yspacing = DY//nr
	calcx = X0 - xspacing/2*(Mc - 2*Mx + 1) + x0 + ptx
	calcy = Y0 + yspacing/2*(Mr - 2*My + 1) + y0 + pty
	shiftx =  - (nc - 2*nx + 1)*(dx - xspacing)/2 - (DX % nc)*(Nc - 2*Nx + 1)/2
	shifty =  + (nr - 2*ny + 1)*(dy - yspacing)/2 + (DY % nr)*(Nr - 2*Ny + 1)/2
	# Kind of a hack move here. Because I moved x0, y0 to the calcx, calcy
	# level, that information from the parent it no longer captured if I only
	# account for shiftx and shifty.
	if parent_args['is_shifted']:
		shiftx += pshiftx
		shifty += pshifty
	#shiftx = np.around(-(Nc - 2*Nx + 1)*(DX - nc*dx)/2, 3)
	#shifty = np.around(+(Nr - 2*Ny + 1)*(DY - nr*dy)/2, 3)
	shiftx = np.around(shiftx, 3)
	shifty = np.around(shifty, 3)
	desiredx = np.around(calcx + shiftx, 3)
	desiredy = np.around(calcy + shifty, 3)

	shiftArray = True
	# I need to be careful of the cases where either dx or dy is zero
	if nc == 1: dx = 0
	if nr == 1: dy = 0

	if dx == 0 and not parent_args['is_shifted']:
		calcx = np.zeros_like(Mx)
		shiftx = np.zeros_like(Mx)
		#desiredx = np.zeros_like(Mx)
		dx = DX
		xspacing = dx

	if dy == 0 and not parent_args['is_shifted']:
		calcy = np.zeros_like(My)
		shifty = np.zeros_like(My)
		#desiredy = np.zeros_like(My)
		dy = DY
		yspacing = dy

	if np.all(shiftx) == 0 or np.all(calcx) == 0:
		calcx = np.zeros_like(Mx)
		shiftx = np.zeros_like(Mx)
		#desiredx = np.zeros_like(Mx)
		DX = dx
		xspacing = dx


	if np.all(shifty) == 0 or np.all(calcy) == 0:
		calcy = np.zeros_like(My)
		shifty = np.zeros_like(My)
		#desiredy = np.zeros_like(My)
		DY = dy
		yspacing = dy

	# Sometimes nested arrays end up falling on a regular grid with exactly no
	# row/column shifts. In this case, we will intuit the existence of this new array and
	# use it.
	if dx == DX and dy == DY and not parent_args['is_shifted']:
		shiftArray = False
		xspacing = dx
		yspacing = dy

	diffx = np.diff(desiredx)
	diffy = np.diff(desiredy)
	xongrid = np.all(diffx == diffx[0]) and nc !=1
	yongrid = np.all(diffy == diffy[0]) and nr != 1
	if xongrid:
		xspacing = np.abs(diffx[0])
		calcx = np.zeros_like(Mx)
		shiftx = np.zeros_like(Mx)
		#desiredx = np.zeros_like(Mx)

	if yongrid:
		yspacing = np.abs(diffy[0])
		calcy = np.zeros_like(My)
		shifty = np.zeros_like(My)
		#desiredy = np.zeros_like(My)

	if xongrid and yongrid:
		shiftArray = False

	# There is another case to be considered here: We could have a global grid
	# with some rows and columns dropped. We also want to check for this
	# possibility because it is easier to program.
	#pdb.set_trace()
	dropRC_args = check_dropped_rows_columns(desiredx, desiredy,
			parent_args['center'])
	if dropRC_args['to_drop']: shiftArray = False
	shift = default
	if shiftArray: shift = mycenter + parent_args['to_shift']
	new_args = {'center':parent_args['center'], 'num_cols':Mc, 'num_rows':Mr,\
			'xspacing':xspacing, 'yspacing':yspacing, 'is_shifted':shiftArray,\
		'column_pos':Mx, 'calculated_x':calcx, 'desired_x':desiredx,\
		'shift_x': shiftx, 'row_pos':My, 'calculated_y':calcy,\
		'desired_y':desiredy, 'shift_y': shifty, 'to_shift':shift,\
		'dropRC_args':dropRC_args}
	return new_args

def makeshot(curr_element, parent_origin=default, parentIsArray=False,
		arrayArgs=empty_dict, mask_list=[], ignored_cells=set()):
	"""
    Function that descends down the hierarchy of cells, starting at the current
    element as the root of the tree and generates a list of all the shots in all 
    the lower levels of the tree and keeps track of the locations of all the
    shots. Each shot represents a unique gdspy CellReference or CellArray
    object and MUST have a single layer.

    Args:
    ---------------
        curr_element (gdspy.Cell): gdspy CellReference object.
        parent_origin (numpy.array): By default numpy.array([0,0]). The origin
            of the element that contains the current element.
        parentIsArray (bool): Boolean value representing whether or not the
            element containing the current element is in an array or not. Required
            in order to correctly place all the subsequent subcells.
            shot.
        arrayArgs (dict): If parentIsArray is True, then this holds all the
            array information.
        mask_list ([gdspy.Cell, ...]): list of all the reticle cells. By
            default, it is the empty list.
            are to be shot.

    Returns:
    ------------------
        ([Shot(), ...]): A list of all the shots in the subtree with the
            current element as the root of the cell hierarchy tree.


    """
	if curr_element.ref_cell.name in ignored_cells: return []
	if type(curr_element) not in allowed_element_types:
		return []
	curr_cell = curr_element.ref_cell
	cell_center = get_center(curr_cell)
	curr_origin = curr_element.origin
	abs_origin = parent_origin + scalearr(curr_origin, scale)
	cell_shift = abs_origin
	cell_size = scalearr(get_size(curr_cell), scale)
	cell_bbox = scalearr(curr_cell.get_bounding_box(), scale)
	haschildren = bool(curr_cell.get_dependencies())

	isArray = False
	dropRC_args = {}
	dropRC_args['to_drop_rows'] = False
	dropRC_args['to_drop_cols'] = False
	if type(curr_element) == gdspy.CellArray and not parentIsArray:
		# Need to correct the array center by the diff btn the dimensions of the ref_cell
		arr_center = get_center(curr_element)
		arr_center = scalearr(arr_center - cell_center , scale) + parent_origin
		abs_origin = default#scalearr(cell_center, scale) # Really crucial. All subcells of the cell shot
		cell_shift = abs_origin
		# in the array have their cell shifts relative to this cell.
		xspacing, yspacing = scalearr(curr_element.spacing, scale)
		newArrayArgs = {'num_cols':curr_element.columns, 'num_rows':curr_element.rows,\
			'center':arr_center, 'xspacing':xspacing, 'yspacing':yspacing,
			'is_shifted':False, 'to_shift':default, 'dropRC_args':dropRC_args}
		isArray = True

	elif type(curr_element) == gdspy.CellReference and not parentIsArray:
		newArrayArgs = {'num_cols':1, 'num_rows':1, 'center':cell_shift,\
				'xspacing':0, 'yspacing':0, 'is_shifted':False,\
				'to_shift':default, 'dropRC_args':dropRC_args}
		cell_shift = default

	elif type(curr_element) == gdspy.CellReference and parentIsArray:
		newArrayArgs = arrayArgs
		isArray = True
		if newArrayArgs['is_shifted']:
			#newArrayArgs['shift_x'] += cell_shift[0]
			#newArrayArgs['shift_y'] += cell_shift[1]
			#newArrayArgs['desired_x'] += cell_shift[0]
			#newArrayArgs['desired_y'] += cell_shift[1]
			if not haschildren:
				cell_shift += newArrayArgs['to_shift']
			else:
				cell_shift = newArrayArgs['to_shift']

	elif type(curr_element) == gdspy.CellArray and parentIsArray:
		abs_origin = default
		arr_center = scalearr(get_center(curr_element), scale)
		#if curr_element.ref_cell.name == "L_MS_feed_bolometer":
		#	print (arr_center)
		arr_center = arr_center - scalearr(cell_center , scale) + parent_origin
		cell_shift = arr_center
		newArrayArgs = get_array_shifts(curr_element, arrayArgs)
		if arrayArgs['is_shifted']: cell_shift = newArrayArgs['to_shift']
		if newArrayArgs['dropRC_args']['to_drop']: cell_shift -= newArrayArgs['dropRC_args']['center']
		isArray = True



	# Sometimes cells have children but are on the mask.  We want to
	# use the largest mask cell possible, so we stop digging early
	# if we find a cell on the mask
	cellonmask = False
	for m in mask_list:
		if curr_cell in m.get_dependencies():
			cellonmask = True

	if cellonmask or (not haschildren):

		try:
			#if curr_element.ref_cell.name == 'ustrip_to_island_R':
			#	pdb.set_trace()
			shot = Shot(curr_cell, cell_shift, cell_size, cell_bbox, isArray=True, **newArrayArgs)
			return [shot]
		except RuntimeError:
			print ("Failed for", curr_element)
			return []

	child_elements = curr_cell.elements
	# If a cell has dependencies, then the elements gives a list of all the cell
	# references in this cell. If the cell has no dependencies, then subelements
	# just gives the polygon set that makes up the cell. I don't want the
	# polygonset stuff.
	child_shotlist = []
	for child in child_elements:
		if type(child) not in allowed_element_types: continue
		# If the current shot is part of a larger array, I want to keep track of
		# that information
		child_shotlist.extend(makeshot(child, abs_origin, isArray, newArrayArgs,\
			mask_list=mask_list, ignored_cells=ignored_cells))
	return child_shotlist
